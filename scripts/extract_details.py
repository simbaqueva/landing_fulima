import openpyxl
import os

excel_file = "Informativa _ Miguel Ramirez _ Informativa basic + pagos personalizados para donaciones _ Miguel Ramirez-subastap.com.co _ undefined _22938938.xlsx"

if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    with open("extracted_info.txt", "w", encoding="utf-8") as f:
        for name in wb.sheetnames:
            f.write(f"\n=========================================\n")
            f.write(f"SHEET: {name}\n")
            f.write(f"=========================================\n")
            sheet = wb[name]
            for r in range(1, sheet.max_row + 1):
                row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
                # check if there is any non-None value in the row
                if any(v is not None for v in row_vals):
                    # format non-None values with their index
                    vals_str = " | ".join([f"Col{i+1}: {v}" for i, v in enumerate(row_vals) if v is not None])
                    f.write(f"Row {r}: {vals_str}\n")
    print("Information extracted to extracted_info.txt successfully!")
else:
    print("Excel file not found!")
