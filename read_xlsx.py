import openpyxl
import pandas as pd
import os

excel_file = "Informativa _ Miguel Ramirez _ Informativa basic + pagos personalizados para donaciones _ Miguel Ramirez-subastap.com.co _ undefined _22938938.xlsx"

if os.path.exists(excel_file):
    print("File exists!")
    # Let's inspect sheet names
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        print(f"\n--- Sheet: {name} ---")
        df = pd.read_excel(excel_file, sheet_name=name)
        print(df.head(20).to_string())
        print(df.info())
else:
    print("File NOT found!")
